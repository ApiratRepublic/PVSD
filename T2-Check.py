"""
T2-Check.py
==================
จุดประสงค์: หาค่ามัธยฐานของราคา ในแต่ละกลุ่มข้อมูล
- Version 1.0.0 0
- release date: 2026-05-11
- ผู้เขียน: Apirat Rattanapaiboon

==========================
"""
from pathlib import Path
from statistics import median
from openpyxl import load_workbook
from openpyxl.styles import Font
from openpyxl.utils import get_column_letter

# =========================================================
# คั้งค่า
# =========================================================
## โฟลเดอรืที่เก็บไฟล์ Excel ทั้งหมด ที่ต้องการประมวลผล (อย่าลืมเปลี่ยนเป็นโฟลเดอร์ของท่าน)
# ผลลลัพธ์จะถูกบันทึกในโฟลเดอร์เดียวกันกับไฟล์ต้นฉบับ โดยมีชื่อไฟล์ที่เพิ่มคำว่า "_with_median" ต่อท้าย
INPUT_DIR = Path(r"D:\xx\xx") 

# ชื่อชีทที่ต้องการประมวลผล (ต้องตรงกับชื่อในไฟล์ Excel)
TARGET_SHEET = "ตาราง 2"

HEADER_ROW = 4 # แถวที่เป็นหัวตาราง (ชื่อคอลัมน์) ในชีทเป้าหมาย (เราจะข้ามแถวนี้ไปนับกลุ่มข้อมูลจริงๆ เริ่มที่แถวถัดไป)
START_ROW = 5 # ข้อมูลเริ่มต้นที่แถวนี้ (นับจาก 1) ในชีทเป้าหมาย

# นับ 1 ที่คอลัมน์ A  ไปจนถึงคอลัมน์ที่ต้องการคำนวณค่ามัธยฐาน
PRICE_COL = 14

# =========================================================
# ฟังก์ชันหลัก
# =========================================================

def to_number(value):
    """
    Convert Excel cell value to float safely.
    """
    if value is None:
        return None

    if isinstance(value, (int, float)):
        return float(value)

    try:
        text = str(value).strip()
        if text == "":
            return None
        text = text.replace(",", "")
        return float(text)
    except:
        return None

def process_sheet(ws):
    max_row = ws.max_row

    median_col = ws.max_column + 1
    ws.cell(HEADER_ROW, median_col).value = "MEDIAN"
    ws.cell(HEADER_ROW, median_col).font = Font(bold=True)

    groups = []
    current_group_start = START_ROW

    for row in range(START_ROW + 1, max_row + 1):
        col_a = ws.cell(row, 1).value
        col_b = ws.cell(row, 2).value

        is_new_group = (
            (col_a is not None and str(col_a).strip() != "") or
            (col_b is not None and str(col_b).strip() != "")
        )

        if is_new_group:
            groups.append((current_group_start, row - 1))
            current_group_start = row

    groups.append((current_group_start, max_row))

    for group_start, group_end in groups:
        values = []
        for r in range(group_start, group_end + 1):
            raw_val = ws.cell(r, PRICE_COL).value
            num_val = to_number(raw_val)

            if num_val is not None:
                values.append(num_val)

        if values:
            med = median(values)
            ws.cell(group_end, median_col).value = med

    ws.column_dimensions[get_column_letter(median_col)].width = 18

    return len(groups)

excel_files = list(INPUT_DIR.glob("*.xlsx"))

if not excel_files:
    print("ไม่พบไฟล์ Excel ใดๆ ในโฟลเดอร์ที่กำหนด. กรุณาตรวจสอบ INPUT_DIR และลองใหม่.")
    exit()

print(f"พบไฟล์ Excel {len(excel_files)} ไฟล์. กำลังเริ่มกระบวนการประมวลผล...")

for file_path in excel_files:

    if file_path.name.startswith("~$"):
        continue

    print(f"\nProcessing: {file_path.name}")

    try:
        wb = load_workbook(file_path, data_only=True)

        if TARGET_SHEET not in wb.sheetnames:
            print(f"  -> ไม่พบชีต '{TARGET_SHEET}' ที่กำหนดไว้ ข้ามไฟล์นี้.")
            continue
        ws = wb[TARGET_SHEET]
        group_count = process_sheet(ws)

        output_name = f"{file_path.stem}_with_median.xlsx"
        output_path = file_path.parent / output_name
        wb.save(output_path)

        print(f"  -> บันทึกสำเร็จ: {output_name}")
        print(f"  -> กลุ่มที่ตรวจพบ: {group_count}")

    except Exception as e:
        print(f"  -> ข้อผิดพลาดขณะประมวลผล {file_path.name}: {e}")

print("\nเสร็จสิ้นการประมวลผลไฟล์ทั้งหมด.")